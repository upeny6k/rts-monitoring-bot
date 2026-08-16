# -*- coding: utf-8 -*-
"""Build RTS Excel from vision extract JSON and move processed images."""
from __future__ import annotations

import json
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"D:\Test Folder\RTS Monitoring project")
INBOX = ROOT / "Yet to be analysed images"
DONE = ROOT / "Analysed Images"
OUT_DIR = ROOT / "Updated report"
TEMPLATE = ROOT / "RTS Report Sample format.xlsx"
EXTRACT_JSON = OUT_DIR / "run_20260804_113624_extract.json"
OUT_XLSX = OUT_DIR / "RTS_2026.08.04_Extracted.xlsx"

DASH = "–"


def safe_move(src: Path, dest_dir: Path) -> Path:
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / src.name
    if not dest.exists():
        shutil.move(str(src), str(dest))
        return dest
    stem, suffix = src.stem, src.suffix
    n = 2
    while True:
        cand = dest_dir / f"{stem}_{n}{suffix}"
        if not cand.exists():
            shutil.move(str(src), str(cand))
            return cand
        n += 1


def main() -> None:
    data = json.loads(EXTRACT_JSON.read_text(encoding="utf-8"))
    parcels = data["parcels"]

    # Sort: corner serial first (all have them here), then whatsapp time
    parcels_sorted = sorted(
        parcels,
        key=lambda p: (
            0 if p.get("corner_serial") is not None else 1,
            p.get("corner_serial") or 0,
            p.get("whatsapp_time") or "",
            p.get("source_image") or "",
        ),
    )

    # Move images first so hyperlinks point to final path
    moved_map: dict[str, Path] = {}
    for name in data.get("images_to_move", []):
        src = INBOX / name
        if src.exists():
            final = safe_move(src, DONE)
            moved_map[name] = final
            print(f"MOVED: {name} -> {final.name}")
        else:
            # already moved?
            alt = DONE / name
            if alt.exists():
                moved_map[name] = alt
                print(f"ALREADY: {name}")
            else:
                print(f"MISSING: {name}")

    wb = load_workbook(TEMPLATE)
    ws = wb.active

    report_date = data.get("report_date", "04.08.2026")
    ws.cell(1, 1).value = (
        f"Monitoring Details of RTS Articles of NSH/PH RMS 'X' DN AGRA, "
        f"DATED:- {report_date}"
    )

    # Ensure extra headers in row 3 (IT 2.0 remark sits next to Col H)
    headers = [
        "SL NO.",
        "Division",
        "Office Name by which Article has been returned",
        "Article No.",
        "Address",
        "Addressee Mobile No.",
        "Whether remark on RTS found Genuine(Yes/No)",
        "If No, Remark",
        "IT 2.0 remark",
        "Source Image (Hyperlink)",
        "AI Confidence",
        "Handwritten RTS Remark",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(3, col)
        cell.value = h
        cell.font = Font(name="Calibri", bold=True, size=11)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    conf_counts = {"high": 0, "medium": 0, "low": 0}

    for i, p in enumerate(parcels_sorted):
        row = 4 + i
        sl = p.get("corner_serial") if p.get("corner_serial") is not None else (i + 1)
        name = (p.get("name") or DASH).strip()
        addr = (p.get("address") or DASH).strip()
        address_combined = f"{name}, {addr}" if name != DASH else addr

        mobile = p.get("mobile") or DASH
        if mobile in ("", "N/A", "n/a", None):
            mobile = DASH

        office = p.get("office_hint") or DASH
        if office in ("", "N/A", "n/a", None):
            office = DASH

        article = p.get("article_no") or DASH
        conf = (p.get("confidence") or "medium").lower()
        if conf not in conf_counts:
            conf = "medium"
        conf_counts[conf] += 1

        remark = p.get("handwritten_remark") or DASH
        if remark in ("", "N/A", "n/a", None):
            remark = DASH

        # Genuine left for phone verification
        genuine = DASH
        if_no = DASH

        values = [
            sl,
            "AGRA",
            office,
            article,
            address_combined,
            mobile,
            genuine,
            if_no,
            DASH,  # IT 2.0 remark — filled by online tracking
            None,  # Source Image hyperlink set below
            conf,
            remark,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row, col)
            cell.value = val
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(name="Calibri", size=10)

        # Hyperlink to final image path (Col J = 10)
        src_name = p.get("source_image") or ""
        img_path = moved_map.get(src_name) or (DONE / src_name)
        link_cell = ws.cell(row, 10)
        if img_path.exists():
            uri = img_path.resolve().as_uri()
            link_cell.value = src_name
            link_cell.hyperlink = uri
            link_cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        else:
            link_cell.value = src_name or DASH

        ws.row_dimensions[row].height = 45

    # Column widths
    widths = {
        1: 8,
        2: 10,
        3: 28,
        4: 18,
        5: 55,
        6: 16,
        7: 18,
        8: 14,
        9: 32,
        10: 42,
        11: 12,
        12: 36,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[3].height = 35

    # Freeze header
    ws.freeze_panes = "A4"

    wb.save(OUT_XLSX)
    print(f"\nExcel saved: {OUT_XLSX}")
    print(f"Parcel rows: {len(parcels_sorted)}")
    print(f"Confidence: high={conf_counts['high']} medium={conf_counts['medium']} low={conf_counts['low']}")
    remaining = list(INBOX.glob("*.*"))
    print(f"Inbox remaining files: {len(remaining)}")


if __name__ == "__main__":
    main()
