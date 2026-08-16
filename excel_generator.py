# -*- coding: utf-8 -*-
"""Excel Generator for RTS Parcel Monitoring Reports."""

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "SL NO.",
    "Division",
    "Office Name by which Article has been returned",
    "Article No.",
    "Address",
    "Addressee Mobile No.",
    "Whether remark on RTS found Genuine(Yes/No)",
    "If No, Remark",
    "IT 2.0 remark",
    "Handwritten RTS Remark",
    "AI Confidence",
    "Source Image",
]


def sort_records(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Sort records: 1st preference = corner_serial (numeric), 2nd = image index."""
    def sort_key(item):
        cs = item.get("corner_serial")
        if cs is not None and isinstance(cs, (int, float)) and cs > 0:
            return (0, int(cs))
        return (1, 999999)

    return sorted(records, key=sort_key)


def build_rts_excel(records: List[Dict[str, Any]], output_path: Path, report_date: str = "") -> Path:
    """Create formatted RTS Excel report with headers, styles, and sorted data."""
    if not report_date:
        report_date = datetime.now().strftime("%d.%m.%Y")

    wb = Workbook()
    ws = wb.active
    ws.title = "RTS Monitoring"

    # Styling Elements
    font_family = "Calibri"
    title_font = Font(name=font_family, size=14, bold=True, color="1F497D")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10)
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Title Block
    ws.merge_cells("A1:L1")
    ws["A1"] = f"POSTAL RTS (RETURN TO SENDER) MONITORING REPORT — {report_date}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Header Row (Row 3)
    ws.row_dimensions[3].height = 28
    for col_idx, header_text in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Sort Records
    sorted_recs = sort_records(records)

    # Write Data Rows (Starting Row 4)
    for row_idx, item in enumerate(sorted_recs, start=4):
        ws.row_dimensions[row_idx].height = 24
        
        # Calculate SL NO.
        cs = item.get("corner_serial")
        sl_val = int(cs) if (cs is not None and str(cs).isdigit()) else (row_idx - 3)

        row_values = [
            sl_val,
            item.get("division", "AGRA"),
            item.get("office", "–"),
            item.get("article_no", "–"),
            item.get("address", "–"),
            item.get("mobile", "–"),
            item.get("genuine", "–"),
            item.get("if_no_remark", "–"),
            item.get("it20_remark", "–"),
            item.get("handwritten_remark", "–"),
            item.get("confidence", "–"),
            item.get("source_image", "–"),
        ]

        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border

            # Alternate row background
            if row_idx % 2 == 1:
                cell.fill = alt_row_fill

            # Alignment
            if col_idx in (1, 2, 4, 6, 7, 8, 11):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in (3,):
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Auto-adjust column widths
    column_widths = {
        1: 8,   # SL NO.
        2: 12,  # Division
        3: 25,  # Office
        4: 18,  # Article No.
        5: 35,  # Address
        6: 18,  # Mobile
        7: 15,  # Genuine
        8: 15,  # If No Remark
        9: 28,  # IT 2.0 remark
        10: 25, # Handwritten RTS Remark
        11: 14, # AI Confidence
        12: 24, # Source Image
    }

    for col_idx, width in column_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
