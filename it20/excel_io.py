# -*- coding: utf-8 -*-
"""Read / write RTS Excel rows for IT 2.0 tracking fill-back.

Column layout (1-based):
  A SL | B Division | C Office (SO from IT 2.0 Destination)
  D Article | E Address | F Mobile | G Genuine | H If No Remark
  I IT 2.0 remark   ← portal delivery/return remark
  J Source Image | K AI Confidence | L Handwritten RTS Remark
  M IT 2.0 Status | N IT 2.0 Tracked At
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from .article_utils import DASH, normalize_article_no

COL = {
    "sl": 1,
    "division": 2,
    "office": 3,
    "article": 4,
    "address": 5,
    "mobile": 6,
    "genuine": 7,
    "if_no_remark": 8,
    "it20_remark": 9,  # title: "IT 2.0 remark"
    "source_image": 10,
    "confidence": 11,
    "handwritten_remark": 12,
    "it20_status": 13,
    "it20_tracked_at": 14,
}

HEADER_ROW = 3
DATA_START = 4

# Canonical headers for extended columns
HEADERS = {
    1: "SL NO.",
    2: "Division",
    3: "Office Name by which Article has been returned",
    4: "Article No.",
    5: "Address",
    6: "Addressee Mobile No.",
    7: "Whether remark on RTS found Genuine(Yes/No)",
    8: "If No, Remark",
    9: "IT 2.0 remark",
    10: "Source Image (Hyperlink)",
    11: "AI Confidence",
    12: "Handwritten RTS Remark",
    13: "IT 2.0 Status",
    14: "IT 2.0 Tracked At",
}


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _find_header_col(ws, *titles: str) -> int | None:
    """Find column index by header title (case-insensitive).

    Prefer exact match, then header startswith title, then title as whole word.
    Avoid naive substring (e.g. \"Article\" must not match \"...Article has been returned\").
    """
    import re

    titles_l = [t.lower().strip() for t in titles]
    # Pass 1: exact
    for c in range(1, (ws.max_column or 1) + 1):
        h = _cell_str(ws.cell(HEADER_ROW, c).value).lower()
        if h in titles_l:
            return c
    # Pass 2: header starts with title or equals ignoring punctuation
    for c in range(1, (ws.max_column or 1) + 1):
        h = _cell_str(ws.cell(HEADER_ROW, c).value).lower()
        if not h:
            continue
        for t in titles_l:
            if h.startswith(t) or h.replace(".", "") == t.replace(".", ""):
                return c
    # Pass 3: whole-word / full-phrase containment only if title has a space or is long
    for c in range(1, (ws.max_column or 1) + 1):
        h = _cell_str(ws.cell(HEADER_ROW, c).value).lower()
        if not h:
            continue
        for t in titles_l:
            if " " in t or len(t) >= 10:
                if t in h:
                    return c
            else:
                if re.search(rf"\b{re.escape(t)}\b", h) and "returned" not in h:
                    return c
    return None


def ensure_tracking_headers(xlsx_path: Path) -> None:
    """
    Ensure IT 2.0 remark sits immediately after Col H.
    Migrates older layouts (Source Image at col 9) by shifting right if needed.
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active

    h9 = _cell_str(ws.cell(HEADER_ROW, 9).value)
    h9_lower = h9.lower()

    # Already correct
    if "it 2.0 remark" in h9_lower:
        for col, title in HEADERS.items():
            if col >= 9 and not _cell_str(ws.cell(HEADER_ROW, col).value):
                cell = ws.cell(HEADER_ROW, col)
                cell.value = title
                cell.font = Font(name="Calibri", bold=True, size=11)
        wb.save(xlsx_path)
        return

    # Old layout: col 9 was Source Image / empty / IT 2.0 Status
    needs_insert = (
        not h9
        or "source" in h9_lower
        or "hyperlink" in h9_lower
        or "status" in h9_lower
        or "delivery" in h9_lower
    )

    if needs_insert and h9 and "it 2.0 remark" not in h9_lower:
        # Shift columns 9..max right by 1
        max_col = max(ws.max_column or 14, 14)
        for r in range(HEADER_ROW, ws.max_row + 1):
            for c in range(max_col, 8, -1):  # from right, stop before H
                ws.cell(r, c + 1).value = ws.cell(r, c).value
                # hyperlinks on source image
                src = ws.cell(r, c)
                dst = ws.cell(r, c + 1)
                if src.hyperlink:
                    dst.hyperlink = src.hyperlink.target if src.hyperlink else None
        # Clear new col 9 data rows (header set below)
        for r in range(DATA_START, ws.max_row + 1):
            ws.cell(r, 9).value = None
            ws.cell(r, 9).hyperlink = None

    for col, title in HEADERS.items():
        cell = ws.cell(HEADER_ROW, col)
        # Don't overwrite core headers 1-8 if already set
        if col <= 8 and cell.value:
            continue
        if col == 9 or not cell.value or col >= 9:
            if col == 9 or not cell.value:
                cell.value = title
            elif col > 9 and title.lower() not in _cell_str(cell.value).lower():
                # Fix known extended titles
                if col in (9, 13, 14) or "source" in title.lower() or "confidence" in title.lower():
                    pass  # keep existing if meaningful
            cell.font = Font(name="Calibri", bold=True, size=11)
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # Force col 9 title
    ws.cell(HEADER_ROW, 9).value = "IT 2.0 remark"
    ws.cell(HEADER_ROW, 9).font = Font(name="Calibri", bold=True, size=11)

    # Ensure trailing headers
    for col in (10, 11, 12, 13, 14):
        if not _cell_str(ws.cell(HEADER_ROW, col).value):
            ws.cell(HEADER_ROW, col).value = HEADERS[col]
            ws.cell(HEADER_ROW, col).font = Font(name="Calibri", bold=True, size=11)

    wb.save(xlsx_path)


def load_parcel_rows(xlsx_path: Path) -> list[dict[str, Any]]:
    """Load data rows from RTS Excel. Skips empty article rows."""
    ensure_tracking_headers(xlsx_path)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    rows: list[dict[str, Any]] = []

    # Resolve columns by header when possible
    col_article = _find_header_col(ws, "Article No.") or COL["article"]
    col_office = (
        _find_header_col(ws, "Office Name by which Article has been returned", "Office Name")
        or COL["office"]
    )
    col_src = _find_header_col(ws, "Source Image (Hyperlink)", "Source Image") or COL["source_image"]
    col_it20 = _find_header_col(ws, "IT 2.0 remark") or COL["it20_remark"]

    for r in range(DATA_START, ws.max_row + 1):
        article_raw = _cell_str(ws.cell(r, col_article).value)
        sl = ws.cell(r, COL["sl"]).value
        if not article_raw and sl is None and not _cell_str(ws.cell(r, COL["address"]).value):
            continue
        if not article_raw:
            continue

        article = normalize_article_no(article_raw)
        rows.append(
            {
                "row": r,
                "sl": sl,
                "division": _cell_str(ws.cell(r, COL["division"]).value) or "AGRA",
                "office": _cell_str(ws.cell(r, col_office).value) or DASH,
                "article": article,
                "article_raw": article_raw,
                "address": _cell_str(ws.cell(r, COL["address"]).value) or DASH,
                "mobile": _cell_str(ws.cell(r, COL["mobile"]).value) or DASH,
                "source_image": _cell_str(ws.cell(r, col_src).value),
                "confidence": _cell_str(ws.cell(r, COL["confidence"]).value),
                "handwritten_remark": _cell_str(ws.cell(r, COL["handwritten_remark"]).value)
                or DASH,
                "it20_remark": _cell_str(ws.cell(r, col_it20).value) or DASH,
            }
        )
    return rows


def update_row_tracking(
    xlsx_path: Path,
    row: int,
    *,
    office: str | None = None,
    article: str | None = None,
    it20_remark: str | None = None,
    status: str | None = None,
    tracked_at: str | None = None,
) -> None:
    """Write tracking results into one Excel row."""
    ensure_tracking_headers(xlsx_path)
    wb = load_workbook(xlsx_path)
    ws = wb.active

    col_office = _find_header_col(ws, "Office Name") or COL["office"]
    col_article = _find_header_col(ws, "Article No.", "Article") or COL["article"]
    col_it20 = _find_header_col(ws, "IT 2.0 remark") or COL["it20_remark"]
    col_status = _find_header_col(ws, "IT 2.0 Status") or COL["it20_status"]
    col_at = _find_header_col(ws, "IT 2.0 Tracked At") or COL["it20_tracked_at"]

    if office is not None and office != "":
        ws.cell(row, col_office).value = office
    if article is not None and article != "":
        ws.cell(row, col_article).value = normalize_article_no(article)
    if it20_remark is not None:
        ws.cell(row, col_it20).value = it20_remark or DASH
    if status is not None:
        ws.cell(row, col_status).value = status or DASH
    if tracked_at is not None:
        ws.cell(row, col_at).value = tracked_at or DASH

    wb.save(xlsx_path)
