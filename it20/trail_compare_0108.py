# -*- coding: utf-8 -*-
"""
Full trail: track all articles in copy Excel, fill Col C + IT 2.0 remark,
then compare against original RTS 01.08.2026.xlsx (manual correct).
"""
from __future__ import annotations

import json
import re
import sys
import time
import unicodedata
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from it20.article_utils import normalize_article_no, is_valid_article_no
from it20.browser import (  # noqa: E402
    login_with_otp,
    open_article_tracking,
    start_session,
    track_article,
)
from it20.excel_io import ensure_tracking_headers, load_parcel_rows, update_row_tracking

COPY = ROOT / "RTS 01.08.2026 - Copy.xlsx"
ORIGINAL = ROOT / "RTS 01.08.2026.xlsx"
STATUS_PATH = ROOT / "Updated report" / "smoke_status.txt"
REPORT_JSON = ROOT / "Updated report" / "trail_01.08.2026_compare.json"
REPORT_TXT = ROOT / "Updated report" / "trail_01.08.2026_compare.txt"


def norm_office(s: str | None) -> str:
    if s is None:
        return ""
    t = str(s).strip()
    t = unicodedata.normalize("NFKC", t)
    t = re.sub(r"\s+", " ", t)
    return t.upper()


def offices_match(auto: str, manual: str) -> bool:
    a, m = norm_office(auto), norm_office(manual)
    if not a or not m or a in ("–", "-") or m in ("–", "-"):
        return False
    if a == m:
        return True
    # soft match: one contains the other (e.g. FATEHABAD vs FATEHABAD SO)
    a2 = re.sub(r"\b(SO|HO|BO|PO|NSH|TMO)\b", "", a).strip()
    m2 = re.sub(r"\b(SO|HO|BO|PO|NSH|TMO)\b", "", m).strip()
    if a2 and m2 and (a2 == m2 or a2 in m or m2 in a):
        return True
    if a in m or m in a:
        return True
    return False


def load_original_map() -> dict[str, dict]:
    wb = load_workbook(ORIGINAL, data_only=True)
    ws = wb.active
    out: dict[str, dict] = {}
    for r in range(4, ws.max_row + 1):
        art = ws.cell(r, 4).value
        if not art:
            continue
        art_n = normalize_article_no(str(art))
        out[art_n] = {
            "sl": ws.cell(r, 1).value,
            "office": ws.cell(r, 3).value,
            "address": ws.cell(r, 5).value,
            "if_no_remark": ws.cell(r, 8).value,
            "row": r,
        }
    return out


def run_track() -> list[dict]:
    ensure_tracking_headers(COPY)
    rows = load_parcel_rows(COPY)
    print(f"Copy Excel parcels: {len(rows)}", flush=True)

    session = start_session(headless=False)
    results: list[dict] = []
    try:
        # Manual TOTP in Chrome; program detects home then starts tracking
        login_with_otp(session, manual_totp=True, manual_timeout_sec=180)
        print("LOGIN_OK", session.page.url, flush=True)
        STATUS_PATH.write_text("LOGIN_OK " + session.page.url, encoding="utf-8")
        open_article_tracking(session)

        for i, row in enumerate(rows, start=1):
            article = normalize_article_no(row["article"])
            print(f"\n[{i}/{len(rows)}] SL={row['sl']}  {article}", flush=True)
            STATUS_PATH.write_text(f"TRACKING {i}/{len(rows)} {article}", encoding="utf-8")

            if not is_valid_article_no(article):
                update_row_tracking(
                    COPY,
                    row["row"],
                    status="invalid_format",
                    it20_remark="–",
                    tracked_at=datetime.now().strftime("%d.%m.%Y %H:%M"),
                )
                results.append(
                    {
                        "sl": row["sl"],
                        "article": article,
                        "excel_row": row["row"],
                        "ok": False,
                        "office": "–",
                        "it20_remark": "–",
                        "error": "invalid_format",
                    }
                )
                continue

            try:
                tr = track_article(session, article)
            except Exception as e:
                print(f"  track exception: {e}", flush=True)
                tr = None
                # re-open tracking page and retry once
                try:
                    open_article_tracking(session)
                    tr = track_article(session, article)
                except Exception as e2:
                    print(f"  retry failed: {e2}", flush=True)

            if tr and tr.ok:
                update_row_tracking(
                    COPY,
                    row["row"],
                    office=tr.office,
                    article=tr.article,
                    it20_remark=tr.it20_remark,
                    status=tr.status,
                    tracked_at=datetime.now().strftime("%d.%m.%Y %H:%M"),
                )
                results.append(
                    {
                        "sl": row["sl"],
                        "article": article,
                        "excel_row": row["row"],
                        "ok": True,
                        "office": tr.office,
                        "it20_remark": tr.it20_remark,
                        "status": tr.status,
                        "error": "",
                    }
                )
            else:
                err = (tr.error if tr else "track failed") or "track failed"
                office = tr.office if tr else "–"
                remark = tr.it20_remark if tr else "–"
                update_row_tracking(
                    COPY,
                    row["row"],
                    office=office if office != "–" else None,
                    it20_remark=remark,
                    status=tr.status if tr else "fail",
                    tracked_at=datetime.now().strftime("%d.%m.%Y %H:%M"),
                )
                results.append(
                    {
                        "sl": row["sl"],
                        "article": article,
                        "excel_row": row["row"],
                        "ok": False,
                        "office": office,
                        "it20_remark": remark,
                        "error": err,
                        "invalid_article": bool(tr.invalid_article) if tr else False,
                    }
                )
            # small pause to avoid hammering portal
            time.sleep(0.4)

        STATUS_PATH.write_text("TRACK_DONE", encoding="utf-8")
        return results
    finally:
        session.close()
        print("Browser closed.", flush=True)


def compare(results: list[dict]) -> dict:
    original = load_original_map()
    exact = soft = mismatch = missing_auto = missing_orig = fail_track = 0
    details = []

    for r in results:
        art = r["article"]
        auto_off = r.get("office") or "–"
        orig = original.get(art)
        if not orig:
            missing_orig += 1
            details.append(
                {
                    **r,
                    "manual_office": None,
                    "match": "not_in_original",
                }
            )
            continue

        manual = orig["office"]
        if not r.get("ok") and (not auto_off or auto_off in ("–", "-")):
            fail_track += 1
            match = "track_failed"
        elif offices_match(str(auto_off), str(manual)):
            if norm_office(str(auto_off)) == norm_office(str(manual)):
                exact += 1
                match = "exact"
            else:
                soft += 1
                match = "soft"
        else:
            if not auto_off or auto_off in ("–", "-"):
                missing_auto += 1
                match = "auto_blank"
            else:
                mismatch += 1
                match = "mismatch"

        details.append(
            {
                "sl": r.get("sl"),
                "article": art,
                "manual_office": manual,
                "auto_office": auto_off,
                "it20_remark": r.get("it20_remark"),
                "match": match,
                "ok": r.get("ok"),
                "error": r.get("error", ""),
            }
        )

    total = len(results)
    compared = exact + soft + mismatch + missing_auto + fail_track
    correct = exact + soft
    summary = {
        "total_tracked_rows": total,
        "original_rows": len(original),
        "exact_match": exact,
        "soft_match": soft,
        "correct_total": correct,
        "mismatch": mismatch,
        "auto_blank": missing_auto,
        "track_failed": fail_track,
        "not_in_original": missing_orig,
        "accuracy_pct_vs_original": round(100.0 * correct / compared, 1) if compared else 0.0,
        "accuracy_exact_pct": round(100.0 * exact / compared, 1) if compared else 0.0,
        "copy_excel": str(COPY),
        "original_excel": str(ORIGINAL),
        "finished_at": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
        "details": details,
    }
    return summary


def write_report(summary: dict) -> None:
    REPORT_JSON.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    lines = [
        "RTS IT 2.0 full trail — 01.08.2026 (re-trail)",
        "Col C rule: SO where article was TO BE DELIVERED",
        "  = Booking Destination (fallback: Item received at Destination office)",
        f"Copy (auto-filled): {COPY.name}",
        f"Original (manual):  {ORIGINAL.name}",
        "Note: original may use a different office concept; compare is reference only.",
        f"Finished: {summary['finished_at']}",
        "",
        f"Total rows tracked:     {summary['total_tracked_rows']}",
        f"Original reference:     {summary['original_rows']}",
        f"Exact Col C match:      {summary['exact_match']}",
        f"Soft Col C match:       {summary['soft_match']}",
        f"Correct (exact+soft):   {summary['correct_total']}",
        f"Mismatch:               {summary['mismatch']}",
        f"Auto blank (no SO):     {summary['auto_blank']}",
        f"Track failed:           {summary['track_failed']}",
        f"Accuracy vs original:   {summary['accuracy_pct_vs_original']}%",
        f"Accuracy exact only:    {summary['accuracy_exact_pct']}%",
        "",
        "--- mismatches / failures ---",
    ]
    for d in summary["details"]:
        if d["match"] in ("exact", "soft"):
            continue
        lines.append(
            f"SL {d.get('sl')} {d.get('article')} | manual={d.get('manual_office')!r} "
            f"auto={d.get('auto_office')!r} | {d.get('match')} {d.get('error') or ''}"
        )
    REPORT_TXT.write_text("\n".join(lines), encoding="utf-8")
    print("\n".join(lines), flush=True)
    print(f"\nJSON: {REPORT_JSON}", flush=True)
    print(f"TXT:  {REPORT_TXT}", flush=True)


def main() -> int:
    if not COPY.exists():
        print(f"Missing copy: {COPY}")
        return 1
    if not ORIGINAL.exists():
        print(f"Missing original: {ORIGINAL}")
        return 1

    print("=== FULL TRAIL 01.08.2026 ===", flush=True)
    print(f"Will fill: {COPY}", flush=True)
    print(f"Compare to: {ORIGINAL}", flush=True)

    results = run_track()
    summary = compare(results)
    write_report(summary)
    STATUS_PATH.write_text("TRAIL_DONE", encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
